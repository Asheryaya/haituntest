import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from models import User, Assessment, DISCResult, DISCAnswer, ScenarioAnswer, SurveyAnswer

def export_users_to_excel(db):
    """导出用户列表到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "用户列表"

    # 表头
    headers = ['ID', '用户名', '姓名', '部门', '角色', '注册时间']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 数据
    users = User.query.all()
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.id)
        ws.cell(row=row, column=2, value=user.username)
        ws.cell(row=row, column=3, value=user.name)
        ws.cell(row=row, column=4, value=user.department or '')
        ws.cell(row=row, column=5, value='管理员' if user.role == 'admin' else '普通用户')
        ws.cell(row=row, column=6, value=user.created_at.strftime('%Y-%m-%d %H:%M'))

    # 调整列宽
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 15

    return wb

def export_disc_results_to_excel(db):
    """导出DISC测评结果到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "DISC测评结果"

    # 表头
    headers = ['用户ID', '姓名', '部门', 'D分数', 'I分数', 'S分数', 'C分数', '主要类型', '次要类型', '完成时间']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 数据
    results = db.session.query(DISCResult, Assessment, User).join(
        Assessment, DISCResult.assessment_id == Assessment.id
    ).join(
        User, Assessment.user_id == User.id
    ).filter(Assessment.status == 'completed').all()

    for row, (result, assessment, user) in enumerate(results, 2):
        ws.cell(row=row, column=1, value=user.id)
        ws.cell(row=row, column=2, value=user.name)
        ws.cell(row=row, column=3, value=user.department or '')
        ws.cell(row=row, column=4, value=result.d_score)
        ws.cell(row=row, column=5, value=result.i_score)
        ws.cell(row=row, column=6, value=result.s_score)
        ws.cell(row=row, column=7, value=result.c_score)
        ws.cell(row=row, column=8, value=result.primary_type)
        ws.cell(row=row, column=9, value=result.secondary_type or '')
        ws.cell(row=row, column=10, value=assessment.completed_at.strftime('%Y-%m-%d %H:%M') if assessment.completed_at else '')

    # 调整列宽
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 15

    return wb

def export_scenario_results_to_excel(db):
    """导出情景测评结果到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "情景测评结果"

    # 表头
    headers = ['用户ID', '姓名', '部门', '情景1-1', '情景1-2', '情景1-3', '情景1-4',
               '情景2-1', '情景2-2', '情景2-3', '情景2-4', '完成时间']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 数据
    assessments = Assessment.query.filter_by(
        assessment_type='scenario', status='completed'
    ).all()

    for row, assessment in enumerate(assessments, 2):
        user = User.query.get(assessment.user_id)
        ws.cell(row=row, column=1, value=user.id)
        ws.cell(row=row, column=2, value=user.name)
        ws.cell(row=row, column=3, value=user.department or '')

        for answer in assessment.scenario_answers:
            ws.cell(row=row, column=3 + answer.scenario_number, value=answer.answer)

        ws.cell(row=row, column=12, value=assessment.completed_at.strftime('%Y-%m-%d %H:%M') if assessment.completed_at else '')

    # 调整列宽
    for col in range(1, 13):
        ws.column_dimensions[chr(64 + col)].width = 15

    return wb

def export_survey_results_to_excel(db):
    """导出培训需求调研结果到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "培训需求调研"

    # 表头
    headers = ['用户ID', '姓名', '部门', '工作亮点/微课课题', '学习期待', '完成时间']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 数据
    assessments = Assessment.query.filter_by(
        assessment_type='survey', status='completed'
    ).all()

    for row, assessment in enumerate(assessments, 2):
        user = User.query.get(assessment.user_id)
        ws.cell(row=row, column=1, value=user.id)
        ws.cell(row=row, column=2, value=user.name)
        ws.cell(row=row, column=3, value=user.department or '')

        for answer in assessment.survey_answers:
            ws.cell(row=row, column=3 + answer.question_number, value=answer.answer)

        ws.cell(row=row, column=6, value=assessment.completed_at.strftime('%Y-%m-%d %H:%M') if assessment.completed_at else '')

    # 调整列宽
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 30

    return wb

def get_statistics(db):
    """获取统计数据"""
    stats = {}

    # 用户统计
    stats['total_users'] = User.query.count()
    stats['admin_users'] = User.query.filter_by(role='admin').count()
    stats['regular_users'] = User.query.filter_by(role='user').count()

    # 测评统计
    stats['total_assessments'] = Assessment.query.count()
    stats['completed_assessments'] = Assessment.query.filter_by(status='completed').count()
    stats['in_progress_assessments'] = Assessment.query.filter_by(status='in_progress').count()

    # 计算未开始测评的用户数（没有任何测评记录的用户）
    users_with_assessments = db.session.query(Assessment.user_id).distinct().count()
    stats['not_started_users'] = stats['total_users'] - users_with_assessments

    # DISC类型分布
    disc_results = DISCResult.query.all()
    disc_distribution = {'D': 0, 'I': 0, 'S': 0, 'C': 0}
    for result in disc_results:
        disc_distribution[result.primary_type] += 1
    stats['disc_distribution'] = disc_distribution

    # 各部分完成率
    disc_completed = Assessment.query.filter_by(assessment_type='disc', status='completed').count()
    scenario_completed = Assessment.query.filter_by(assessment_type='scenario', status='completed').count()
    survey_completed = Assessment.query.filter_by(assessment_type='survey', status='completed').count()

    stats['disc_completion_rate'] = (disc_completed / stats['total_users'] * 100) if stats['total_users'] > 0 else 0
    stats['scenario_completion_rate'] = (scenario_completed / stats['total_users'] * 100) if stats['total_users'] > 0 else 0
    stats['survey_completion_rate'] = (survey_completed / stats['total_users'] * 100) if stats['total_users'] > 0 else 0

    return stats
